"""查驗報告匯出：Excel (.xlsx) 與 PDF。

提供承辦人可列印、可歸檔的正式查驗報告，內容含任務資訊、風險統計、
逐筆遮罩後發現，以及審核紀錄（誰、何時、什麼決定）。
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

from .db import get_db, row_to_dict
from .usage_control import job_usage

RISK_LABELS = {"High": "高風險", "Medium": "中風險", "Low": "低風險", "None": "無"}
DECISION_LABELS = {
    "approved": "通過",
    "needs_changes": "需修改",
    "false_positive": "誤判",
    "approved_after_redaction": "遮罩後通過",
}


def collect_report(job_id: str) -> dict[str, Any] | None:
    """彙整單一任務的完整報告資料；查無任務時回傳 None。"""
    db = get_db()
    job = row_to_dict(db.execute("SELECT * FROM scan_jobs WHERE id = ?", (job_id,)).fetchone())
    if not job:
        return None
    findings = [
        dict(row)
        for row in db.execute(
            """
            SELECT files.original_name, findings.category, findings.risk_level,
                   findings.confidence, findings.masked_text, findings.location,
                   findings.recommendation, findings.detector
            FROM findings
            JOIN files ON files.id = findings.file_id
            WHERE findings.job_id = ?
            ORDER BY
              CASE findings.risk_level WHEN 'High' THEN 1 WHEN 'Medium' THEN 2
                   WHEN 'Low' THEN 3 ELSE 4 END,
              files.original_name
            """,
            (job_id,),
        ).fetchall()
    ]
    files = [
        dict(row)
        for row in db.execute(
            "SELECT original_name, extension, size, status, error FROM files WHERE job_id = ?",
            (job_id,),
        ).fetchall()
    ]
    reviews = [
        dict(row)
        for row in db.execute(
            "SELECT decision, reviewer, note, created_at FROM reviews WHERE job_id = ? ORDER BY created_at",
            (job_id,),
        ).fetchall()
    ]
    return {
        "job": job,
        "files": files,
        "findings": findings,
        "reviews": reviews,
        "aiUsage": job_usage(job_id),
        "generatedAt": datetime.now(timezone.utc).isoformat(),
    }


def _risk_counts(findings: list[dict[str, Any]]) -> dict[str, int]:
    counts = {"High": 0, "Medium": 0, "Low": 0}
    for finding in findings:
        level = finding.get("risk_level")
        if level in counts:
            counts[level] += 1
    return counts


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def build_excel(report: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    job = report["job"]
    findings = report["findings"]
    counts = _risk_counts(findings)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    # 摘要工作表
    summary = wb.active
    summary.title = "查驗摘要"
    summary_rows = [
        ("項目", "內容"),
        ("任務編號", job.get("id", "")),
        ("整體風險", RISK_LABELS.get(job.get("risk_level"), job.get("risk_level", ""))),
        ("狀態", job.get("status", "")),
        ("建立時間", job.get("created_at", "")),
        ("完成時間", job.get("completed_at", "")),
        ("高風險筆數", counts["High"]),
        ("中風險筆數", counts["Medium"]),
        ("低風險筆數", counts["Low"]),
        ("發現總數", len(findings)),
        ("報告產生時間", report.get("generatedAt", "")),
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 48

    # 風險發現工作表
    ws = wb.create_sheet("風險發現")
    headers = ["風險等級", "檔案", "類型", "信心度", "遮罩內容", "位置", "建議"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for finding in findings:
        ws.append(
            [
                RISK_LABELS.get(finding.get("risk_level"), finding.get("risk_level", "")),
                finding.get("original_name", ""),
                finding.get("category", ""),
                round(float(finding.get("confidence") or 0), 2),
                finding.get("masked_text", ""),
                finding.get("location", ""),
                finding.get("recommendation", ""),
            ]
        )
    widths = [12, 24, 16, 10, 24, 28, 40]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 審核紀錄工作表
    rv = wb.create_sheet("審核紀錄")
    rv.append(["審核決定", "審核人", "備註", "時間"])
    for cell in rv[1]:
        cell.fill = header_fill
        cell.font = header_font
    for review in report["reviews"]:
        rv.append(
            [
                DECISION_LABELS.get(review.get("decision"), review.get("decision", "")),
                review.get("reviewer", ""),
                review.get("note", ""),
                review.get("created_at", ""),
            ]
        )
    for index, width in enumerate([16, 24, 40, 26], start=1):
        rv.column_dimensions[get_column_letter(index)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# PDF
# --------------------------------------------------------------------------- #
_FONT_NAME = "STSong-Light"
_font_registered = False


def _ensure_font() -> str:
    """註冊內建 Adobe CJK 字型（免外部字型檔，支援中文顯示）。"""
    global _font_registered
    if not _font_registered:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont

        pdfmetrics.registerFont(UnicodeCIDFont(_FONT_NAME))
        _font_registered = True
    return _FONT_NAME


def build_pdf(report: dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font = _ensure_font()
    job = report["job"]
    findings = report["findings"]
    counts = _risk_counts(findings)

    base = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=base["Title"], fontName=font, fontSize=18)
    h2_style = ParagraphStyle("h2", parent=base["Heading2"], fontName=font, fontSize=13)
    body_style = ParagraphStyle("body", parent=base["Normal"], fontName=font, fontSize=9, leading=12)
    note_style = ParagraphStyle("note", parent=body_style, textColor=colors.grey, fontSize=8)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=16 * mm,
        title="個資查驗報告",
    )
    story: list[Any] = [Paragraph("個資查驗報告", title_style), Spacer(1, 6 * mm)]

    def cell(text: Any) -> Paragraph:
        return Paragraph("" if text is None else str(text), body_style)

    # 摘要
    story.append(Paragraph("一、任務資訊", h2_style))
    summary_data = [
        [cell("任務編號"), cell(job.get("id", ""))],
        [cell("整體風險"), cell(RISK_LABELS.get(job.get("risk_level"), job.get("risk_level", "")))],
        [cell("狀態"), cell(job.get("status", ""))],
        [cell("建立時間"), cell(job.get("created_at", ""))],
        [cell("完成時間"), cell(job.get("completed_at", ""))],
        [
            cell("風險統計"),
            cell(f"高風險 {counts['High']}　中風險 {counts['Medium']}　低風險 {counts['Low']}　共 {len(findings)} 筆"),
        ],
    ]
    summary_table = Table(summary_data, colWidths=[30 * mm, 148 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f3f8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 6 * mm))

    # 風險發現
    story.append(Paragraph("二、風險發現（遮罩後）", h2_style))
    if findings:
        head = ["風險", "檔案", "類型", "遮罩內容", "位置", "建議"]
        table_data = [[cell(h) for h in head]]
        for finding in findings:
            table_data.append(
                [
                    cell(RISK_LABELS.get(finding.get("risk_level"), finding.get("risk_level", ""))),
                    cell(finding.get("original_name", "")),
                    cell(finding.get("category", "")),
                    cell(finding.get("masked_text", "")),
                    cell(finding.get("location", "")),
                    cell(finding.get("recommendation", "")),
                ]
            )
        findings_table = Table(
            table_data,
            colWidths=[14 * mm, 28 * mm, 24 * mm, 30 * mm, 30 * mm, 52 * mm],
            repeatRows=1,
        )
        findings_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(findings_table)
    else:
        story.append(Paragraph("未發現疑似個資。", body_style))
    story.append(Spacer(1, 6 * mm))

    # 審核紀錄
    story.append(Paragraph("三、審核紀錄", h2_style))
    if report["reviews"]:
        head = ["審核決定", "審核人", "備註", "時間"]
        review_data = [[cell(h) for h in head]]
        for review in report["reviews"]:
            review_data.append(
                [
                    cell(DECISION_LABELS.get(review.get("decision"), review.get("decision", ""))),
                    cell(review.get("reviewer", "")),
                    cell(review.get("note", "")),
                    cell(review.get("created_at", "")),
                ]
            )
        review_table = Table(review_data, colWidths=[24 * mm, 34 * mm, 70 * mm, 50 * mm], repeatRows=1)
        review_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(review_table)
    else:
        story.append(Paragraph("尚無審核紀錄。", body_style))

    story.append(Spacer(1, 8 * mm))
    story.append(
        Paragraph(
            f"報告產生時間：{report.get('generatedAt', '')}　"
            "本系統僅為輔助查驗工具，最終結果仍須由承辦人員確認。",
            note_style,
        )
    )

    doc.build(story)
    return buffer.getvalue()
